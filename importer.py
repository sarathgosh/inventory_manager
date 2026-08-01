from __future__ import annotations

import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SOURCE_SHEET = "Inventory List"


def _text(value: Any) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return text or None


def _number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        try:
            return None if pd.isna(value) else float(value)
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    if text.upper() in {"FREE", "NO COST"}:
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", text.replace(",", ""))
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _integer(value: Any) -> int | None:
    number = _number(value)
    return int(number) if number is not None else None


def _date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 80000:
        try:
            parsed = from_excel(float(value))
            return parsed.date().isoformat() if isinstance(parsed, datetime) else parsed.isoformat()
        except (ValueError, OverflowError):
            return None
    text = str(value).strip()
    if not text or text in {"00:00:00", "0:00:00"}:
        return None
    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%y",
        "%d/%m/%y",
    ):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _source_note(label: str, value: Any) -> str | None:
    text = _text(value)
    return f"Unparsed {label}: {text}" if text else None


def parse_source_workbook(path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parse the user's original workbook without changing it."""
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook[SOURCE_SHEET] if SOURCE_SHEET in workbook.sheetnames else workbook.active
    records: list[dict[str, Any]] = []
    invalid_opened_dates = 0
    invalid_expiry_dates = 0
    invalid_prices = 0

    # The source workbook uses row 3 as its table header. The meaningful data
    # fields occupy columns C:X, while column H contains the required item name.
    for row_number in range(4, sheet.max_row + 1):
        item_name = _text(sheet.cell(row_number, 8).value)
        if not item_name:
            continue
        notes: list[str] = []

        unit_price_source = sheet.cell(row_number, 12).value
        unit_price = _number(unit_price_source)
        if unit_price is None and _text(unit_price_source):
            invalid_prices += 1
            notes.append(_source_note("unit price", unit_price_source) or "")
        elif _text(unit_price_source) and str(unit_price_source).strip().upper() == "FREE":
            notes.append("Source unit price was marked FREE")

        opened_source = sheet.cell(row_number, 20).value
        opened_date = _date(opened_source)
        if opened_date is None and _text(opened_source):
            invalid_opened_dates += 1
            notes.append(_source_note("opened date", opened_source) or "")

        expiry_source = sheet.cell(row_number, 21).value
        expiry_date = _date(expiry_source)
        if expiry_date is None and _text(expiry_source):
            invalid_expiry_dates += 1
            notes.append(_source_note("expiry date", expiry_source) or "")

        reorder_due_source = sheet.cell(row_number, 23).value
        reorder_due = _date(reorder_due_source)
        if reorder_due is None and _text(reorder_due_source):
            notes.append(_source_note("reorder due date", reorder_due_source) or "")

        reminder_days = _integer(sheet.cell(row_number, 22).value)
        if reorder_due is None and opened_date and reminder_days is not None:
            reorder_due = (
                pd.Timestamp(opened_date) + pd.Timedelta(days=reminder_days)
            ).date().isoformat()

        record = {
            "inventory_code": None,
            "item_name": item_name,
            "brand": _text(sheet.cell(row_number, 9).value),
            "category": "Laboratory Inventory",
            "test_volume": _text(sheet.cell(row_number, 10).value),
            "description": _text(sheet.cell(row_number, 11).value),
            "unit_price": unit_price,
            "total_test_effective": _number(sheet.cell(row_number, 13).value),
            "selling_price_per_test": _number(sheet.cell(row_number, 16).value),
            "quantity_in_stock": _number(sheet.cell(row_number, 18).value),
            "reorder_level": _number(sheet.cell(row_number, 3).value),
            "reorder_quantity": _number(sheet.cell(row_number, 24).value),
            "opened_date": opened_date,
            "expiry_date": expiry_date,
            "reminder_days": reminder_days,
            "reorder_due_date": reorder_due,
            "batch_no": None,
            "supplier": None,
            "storage_location": None,
            "notes": None,
            "source_inventory_id": _text(sheet.cell(row_number, 4).value),
            "source_row": row_number,
            "source_notes": "; ".join(note for note in notes if note) or None,
            "is_active": 1,
        }
        records.append(record)

    diagnostics = {
        "sheet": sheet.title,
        "records": len(records),
        "invalid_opened_dates": invalid_opened_dates,
        "invalid_expiry_dates": invalid_expiry_dates,
        "invalid_prices": invalid_prices,
        "records_missing_stock": sum(
            1 for record in records if record["quantity_in_stock"] is None
        ),
        "records_missing_expiry": sum(1 for record in records if record["expiry_date"] is None),
    }
    return records, diagnostics


ALIASES = {
    "inventory_code": {"inventory code", "inventory id", "code"},
    "item_name": {"item", "item name", "product", "product name"},
    "brand": {"brand", "manufacturer"},
    "category": {"category", "item category"},
    "test_volume": {"test volume", "volume", "test/volume", "pack volume"},
    "description": {"description", "pack size", "details"},
    "unit_price": {"unit price", "purchase price", "cost price"},
    "total_test_effective": {"total test effective", "total tests", "effective tests"},
    "selling_price_per_test": {"selling price per test", "selling price", "sale price per test"},
    "quantity_in_stock": {"quantity in stock", "stock quantity", "quantity", "current stock"},
    "reorder_level": {"reorder level", "for reorder", "minimum stock", "min stock"},
    "reorder_quantity": {"quantity in reorder", "reorder quantity", "quantity to reorder"},
    "opened_date": {"opened date", "date opened"},
    "expiry_date": {"expiry date", "expiration date", "expiry"},
    "reminder_days": {"reminder in days", "reminder days"},
    "reorder_due_date": {"re-order due", "reorder due", "reorder due date"},
    "batch_no": {"batch no", "batch number", "lot no", "lot number", "batch / lot no."},
    "supplier": {"supplier", "vendor"},
    "storage_location": {"storage location", "location", "rack"},
    "notes": {"notes", "remarks", "comment"},
    "source_inventory_id": {"source inventory id"},
}


def _normalise_header(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip().lower()
    return text


def _column_mapping(columns: Iterable[Any]) -> dict[str, Any]:
    mapping: dict[str, Any] = {}
    normalised = {_normalise_header(column): column for column in columns}
    for target, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                mapping[target] = normalised[alias]
                break
    return mapping


def _read_tabular_bytes(filename: str, content: bytes) -> pd.DataFrame:
    lower_name = filename.lower()
    buffer = io.BytesIO(content)
    if lower_name.endswith(".csv"):
        return pd.read_csv(buffer)
    if not lower_name.endswith((".xlsx", ".xlsm")):
        raise ValueError("Upload an .xlsx, .xlsm, or .csv file.")
    excel = pd.ExcelFile(buffer, engine="openpyxl")
    sheet_name = SOURCE_SHEET if SOURCE_SHEET in excel.sheet_names else excel.sheet_names[0]
    preview = pd.read_excel(excel, sheet_name=sheet_name, header=None, nrows=15)
    header_index = 0
    for index, row in preview.iterrows():
        values = {_normalise_header(value) for value in row.tolist()}
        if "item" in values or "item name" in values:
            header_index = int(index)
            break
    return pd.read_excel(excel, sheet_name=sheet_name, header=header_index)


def parse_tabular_upload(filename: str, content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    frame = _read_tabular_bytes(filename, content)
    frame = frame.dropna(axis=1, how="all")
    mapping = _column_mapping(frame.columns)
    if "item_name" not in mapping:
        raise ValueError("Could not find an Item or Item Name column in the uploaded file.")

    records: list[dict[str, Any]] = []
    for source_index, row in frame.iterrows():
        item_name = _text(row.get(mapping["item_name"]))
        if not item_name:
            continue
        notes: list[str] = []

        def source(field: str) -> Any:
            column = mapping.get(field)
            return row.get(column) if column is not None else None

        opened_date = _date(source("opened_date"))
        expiry_date = _date(source("expiry_date"))
        reorder_due = _date(source("reorder_due_date"))
        if opened_date is None and _text(source("opened_date")):
            notes.append(_source_note("opened date", source("opened_date")) or "")
        if expiry_date is None and _text(source("expiry_date")):
            notes.append(_source_note("expiry date", source("expiry_date")) or "")

        record = {
            "inventory_code": _text(source("inventory_code")),
            "item_name": item_name,
            "brand": _text(source("brand")),
            "category": _text(source("category")) or "Laboratory Inventory",
            "test_volume": _text(source("test_volume")),
            "description": _text(source("description")),
            "unit_price": _number(source("unit_price")),
            "total_test_effective": _number(source("total_test_effective")),
            "selling_price_per_test": _number(source("selling_price_per_test")),
            "quantity_in_stock": _number(source("quantity_in_stock")),
            "reorder_level": _number(source("reorder_level")),
            "reorder_quantity": _number(source("reorder_quantity")),
            "opened_date": opened_date,
            "expiry_date": expiry_date,
            "reminder_days": _integer(source("reminder_days")),
            "reorder_due_date": reorder_due,
            "batch_no": _text(source("batch_no")),
            "supplier": _text(source("supplier")),
            "storage_location": _text(source("storage_location")),
            "notes": _text(source("notes")),
            "source_inventory_id": _text(source("source_inventory_id")),
            "source_row": int(source_index) + 2,
            "source_notes": "; ".join(note for note in notes if note) or None,
            "is_active": 1,
        }
        records.append(record)

    diagnostics = {
        "records": len(records),
        "mapped_fields": sorted(mapping),
        "unmapped_columns": [
            str(column) for column in frame.columns if column not in set(mapping.values())
        ],
        "records_missing_stock": sum(
            1 for record in records if record["quantity_in_stock"] is None
        ),
        "records_missing_expiry": sum(1 for record in records if record["expiry_date"] is None),
    }
    return records, diagnostics


def records_preview(records: list[Mapping[str, Any]], limit: int = 100) -> pd.DataFrame:
    columns = [
        "inventory_code",
        "item_name",
        "brand",
        "description",
        "unit_price",
        "quantity_in_stock",
        "reorder_level",
        "expiry_date",
        "source_notes",
    ]
    return pd.DataFrame(records[:limit]).reindex(columns=columns)
