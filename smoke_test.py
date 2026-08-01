from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
import sqlite3
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from catalog import database_options, filter_by_option, latest_record, numeric_options
from database import (
    add_item,
    connect,
    delete_item,
    fetch_inventory_dataframe,
    fetch_movements_dataframe,
    get_item,
    initialize_database,
    insert_items,
    make_inventory_code,
    recode_all_inventory,
    record_stock_movement,
    restore_item,
    update_item,
)
from importer import parse_source_workbook, parse_tabular_upload
from reports import (
    build_report_frame,
    create_excel_report,
    create_import_template,
    inventory_summary,
)


def main() -> None:
    app_dir = Path(__file__).resolve().parent
    seed = app_dir / "data" / "Inventory_list_source.xlsx"
    assert seed.exists(), "Seed workbook is missing"

    records, diagnostics = parse_source_workbook(seed)
    assert diagnostics["records"] == 407, diagnostics
    assert len(records) == 407

    with TemporaryDirectory() as temporary_directory:
        legacy_database_path = Path(temporary_directory) / "legacy_inventory.db"
        with sqlite3.connect(legacy_database_path) as legacy_connection:
            legacy_connection.executescript(
                """
                CREATE TABLE inventory (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    inventory_code TEXT NOT NULL UNIQUE,
                    item_name TEXT NOT NULL,
                    brand TEXT,
                    description TEXT,
                    expiry_date TEXT,
                    is_active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    archived_at TEXT
                );
                INSERT INTO inventory(
                    inventory_code, item_name, brand, description, is_active,
                    created_at, archived_at
                ) VALUES (
                    'LEGACY-1', 'LEGACY ITEM', 'TEST', '1 X 1 ML', 0,
                    '2025-01-01 10:00:00', '2025-02-01 11:00:00'
                );
                """
            )
        initialize_database(legacy_database_path)
        legacy = get_item(1, legacy_database_path)
        assert legacy["create_date"] is None
        assert legacy["total_test_available"] is None
        assert legacy["total_test_done"] is None
        assert legacy["finish_date"] is None
        assert legacy["deleted_at"] == "2025-02-01 11:00:00"

        database_path = Path(temporary_directory) / "inventory_test.db"
        initialize_database(database_path)
        result = insert_items(records, replace=True, db_path=database_path)
        assert result == {"inserted": 407, "skipped": 0}, result

        inventory = fetch_inventory_dataframe(db_path=database_path, warning_days=30)
        assert len(inventory) == 407
        assert inventory["inventory_code"].is_unique
        assert sorted(inventory["id"].tolist()) == list(range(1, 408))
        assert all(
            row.inventory_code == make_inventory_code(row.item_name, row.description, row.id)
            for row in inventory.itertuples()
        )
        assert inventory["source_row"].min() == 4
        assert inventory["source_row"].max() == 410
        assert inventory["create_date"].isna().all()
        assert inventory["total_test_available"].isna().all()
        assert inventory["total_test_done"].isna().all()
        assert inventory["finish_date"].isna().all()

        item_options = database_options(inventory, "item_name", include_missing=False)
        assert len(item_options) == 29
        selected_item = "SGOT"
        item_rows = filter_by_option(inventory, "item_name", selected_item)
        assert not item_rows.empty
        brand_options = database_options(item_rows, "brand")
        brand_rows = filter_by_option(item_rows, "brand", brand_options[0])
        assert not brand_rows.empty
        assert latest_record(brand_rows) is not None
        assert numeric_options(brand_rows, "unit_price")

        with connect(database_path) as connection:
            connection.execute(
                "UPDATE inventory SET inventory_code = printf('INV-%05d', id)"
            )
            connection.commit()
        assert recode_all_inventory(database_path) == 407
        migrated = get_item(1, database_path)
        assert migrated["inventory_code"] == make_inventory_code(
            migrated["item_name"], migrated["description"], 1
        )

        test_item_id = add_item(
            {
                "item_name": "SMOKE TEST ITEM",
                "description": "2 x 30 ML",
                "brand": "TEST",
                "quantity_in_stock": None,
                "total_test_available": 30,
                "total_test_done": 5,
                "finish_date": "2030-11-30",
                "reorder_level": 2,
                "unit_price": 10,
                "expiry_date": "2030-12-31",
            },
            db_path=database_path,
        )
        assert test_item_id == 408
        created = get_item(test_item_id, database_path)
        assert created["inventory_code"] == "STI-2X30ML-00408"
        assert created["create_date"] == date.today().isoformat()
        assert created["total_test_available"] == 30
        assert created["total_test_done"] == 5
        assert created["finish_date"] == "2030-11-30"
        update_item(
            test_item_id,
            {
                "description": "5X100",
                "total_test_available": 24,
                "total_test_done": 11,
                "finish_date": "2030-12-15",
                "create_date": "1999-01-01",
            },
            database_path,
        )
        created = get_item(test_item_id, database_path)
        assert created["inventory_code"] == "STI-5X100-00408"
        assert created["create_date"] == date.today().isoformat()
        assert created["total_test_available"] == 24
        assert created["total_test_done"] == 11
        assert created["finish_date"] == "2030-12-15"
        balance = record_stock_movement(
            test_item_id,
            "Receive",
            5,
            reference="TEST-RECEIPT",
            db_path=database_path,
        )
        assert balance == 5
        balance = record_stock_movement(
            test_item_id,
            "Issue",
            2,
            reference="TEST-ISSUE",
            db_path=database_path,
        )
        assert balance == 3
        assert get_item(test_item_id, database_path)["quantity_in_stock"] == 3
        assert len(fetch_movements_dataframe(database_path)) == 2

        delete_item(test_item_id, database_path)
        deleted_record = get_item(test_item_id, database_path)
        assert deleted_record["is_active"] == 0
        assert deleted_record["deleted_at"]
        active_after_delete = fetch_inventory_dataframe(
            db_path=database_path, warning_days=30
        )
        assert len(active_after_delete) == 407
        deleted = fetch_inventory_dataframe(
            deleted_only=True, db_path=database_path, warning_days=30
        )
        assert len(deleted) == 1
        assert deleted.iloc[0]["overall_status"] == "Deleted"
        deleted_report = build_report_frame("Deleted Inventory Report", deleted)
        assert len(deleted_report) == 1
        assert "deleted_at" in deleted_report.columns
        assert "total_test_available" in deleted_report.columns
        assert "total_test_done" in deleted_report.columns
        assert "finish_date" in deleted_report.columns

        restore_item(test_item_id, database_path)
        assert get_item(test_item_id, database_path)["is_active"] == 1
        assert fetch_inventory_dataframe(
            deleted_only=True, db_path=database_path, warning_days=30
        ).empty

        inventory = fetch_inventory_dataframe(db_path=database_path, warning_days=30)
        report = build_report_frame("Full Inventory Report", inventory)
        report_bytes = create_excel_report(
            "Full Inventory Report",
            report,
            inventory_summary(inventory),
            "₹",
        )
        workbook = load_workbook(BytesIO(report_bytes), read_only=True)
        assert workbook.sheetnames == ["Summary", "Report Data"]
        assert workbook["Report Data"].max_row == len(report) + 1
        assert workbook["Report Data"]["A1"].value == "Inventory ID"
        report_headers = [
            cell.value for cell in next(workbook["Report Data"].iter_rows(min_row=1, max_row=1))
        ]
        for expected_header in (
            "Create Date",
            "Total Test Avail",
            "Total Test Done",
            "Finish Date",
        ):
            assert expected_header in report_headers

        template = load_workbook(BytesIO(create_import_template()), read_only=True)
        template_headers = [
            cell.value
            for cell in next(template["Inventory Import"].iter_rows(min_row=1, max_row=1))
        ]
        for expected_header in (
            "Create Date",
            "Total Test Avail",
            "Total Test Done",
            "Finish Date",
        ):
            assert expected_header in template_headers

        imported_rows, imported_diagnostics = parse_tabular_upload(
            "test_tracking.csv",
            (
                "Item Name,Create Date,Total Test Avail,Total Test Done,Finish Date\n"
                "CSV TEST,2031-01-02,50,12,2031-02-03\n"
            ).encode("utf-8"),
        )
        assert imported_diagnostics["records"] == 1
        assert imported_rows[0]["create_date"] == "2031-01-02"
        assert imported_rows[0]["total_test_available"] == 50
        assert imported_rows[0]["total_test_done"] == 12
        assert imported_rows[0]["finish_date"] == "2031-02-03"

    print(
        "Smoke test passed: 407 source records imported; CRUD, cascading database options, "
        "incremental IDs, code migration, stock movements, "
        "blank legacy/automatic new create dates, test tracking, soft delete/restore, "
        "calculated statuses and Excel reporting verified."
    )


if __name__ == "__main__":
    main()
