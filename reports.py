from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAVY = "12304A"
TEAL = "00A6A6"
LIGHT_TEAL = "DDF4F4"
LIGHT_BLUE = "EAF1F7"
RED = "D64545"
AMBER = "F4B942"
GREEN = "2D8C6A"
WHITE = "FFFFFF"
GREY = "637381"


FRIENDLY_COLUMNS = {
    "id": "Inventory ID",
    "inventory_id": "Inventory ID",
    "inventory_code": "Inventory Code",
    "item_name": "Item Name",
    "brand": "Brand",
    "category": "Category",
    "test_volume": "Test Volume",
    "description": "Description",
    "unit_price": "Unit Price",
    "total_test_effective": "Total Test Effective",
    "total_test_available": "Total Test Avail",
    "total_test_done": "Total Test Done",
    "cost_per_test": "Cost Per Test",
    "selling_price_per_test": "Selling Price Per Test",
    "profit_per_test": "Profit Per Test",
    "quantity_in_stock": "Quantity In Stock",
    "inventory_value": "Inventory Value",
    "reorder_level": "Reorder Level",
    "reorder_quantity": "Reorder Quantity",
    "create_date": "Create Date",
    "opened_date": "Opened Date",
    "expiry_date": "Expiry Date",
    "finish_date": "Finish Date",
    "days_to_expiry": "Days To Expiry",
    "reminder_days": "Reminder Days",
    "reorder_due_date": "Reorder Due Date",
    "batch_no": "Batch / Lot No.",
    "supplier": "Supplier",
    "storage_location": "Storage Location",
    "stock_status": "Stock Status",
    "expiry_status": "Expiry Status",
    "overall_status": "Overall Status",
    "data_quality_issues": "Data Quality Issues",
    "notes": "Notes",
    "source_inventory_id": "Source Inventory ID",
    "source_row": "Source Row",
    "source_notes": "Source Notes",
    "is_active": "Active",
    "moved_at": "Movement Date",
    "movement_type": "Movement Type",
    "quantity_change": "Quantity Change",
    "quantity_before": "Quantity Before",
    "quantity_after": "Quantity After",
    "reference": "Reference",
    "created_at": "Recorded At",
    "deleted_at": "Deleted At",
}


INVENTORY_REPORT_COLUMNS = [
    "id",
    "inventory_code",
    "item_name",
    "brand",
    "category",
    "test_volume",
    "description",
    "batch_no",
    "create_date",
    "unit_price",
    "total_test_effective",
    "total_test_available",
    "total_test_done",
    "cost_per_test",
    "selling_price_per_test",
    "profit_per_test",
    "quantity_in_stock",
    "inventory_value",
    "reorder_level",
    "reorder_quantity",
    "opened_date",
    "expiry_date",
    "finish_date",
    "days_to_expiry",
    "reorder_due_date",
    "supplier",
    "storage_location",
    "stock_status",
    "expiry_status",
    "overall_status",
    "data_quality_issues",
    "notes",
    "source_notes",
]


DELETED_REPORT_COLUMNS = [*INVENTORY_REPORT_COLUMNS, "deleted_at"]


MOVEMENT_REPORT_COLUMNS = [
    "moved_at",
    "inventory_id",
    "inventory_code",
    "item_name",
    "brand",
    "movement_type",
    "quantity_change",
    "quantity_before",
    "quantity_after",
    "reference",
    "notes",
    "created_at",
]


def inventory_summary(frame: pd.DataFrame) -> dict[str, Any]:
    if frame.empty:
        return {
            "Total Records": 0,
            "Unique Items": 0,
            "Known Inventory Value": 0.0,
            "Reorder Alerts": 0,
            "Expired Records": 0,
            "Expiring Soon": 0,
            "Stock Quantity Missing": 0,
        }
    return {
        "Total Records": int(len(frame)),
        "Unique Items": int(frame["item_name"].nunique(dropna=True)),
        "Known Inventory Value": float(frame["inventory_value"].sum(skipna=True)),
        "Reorder Alerts": int(frame["stock_status"].isin(["Out of Stock", "Low Stock"]).sum()),
        "Expired Records": int(frame["expiry_status"].eq("Expired").sum()),
        "Expiring Soon": int(frame["expiry_status"].eq("Expiring Soon").sum()),
        "Stock Quantity Missing": int(frame["quantity_in_stock"].isna().sum()),
    }


def build_report_frame(
    report_name: str,
    inventory: pd.DataFrame,
    movements: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if report_name == "Stock Movement Report":
        movement_frame = movements if movements is not None else pd.DataFrame()
        return _select_columns(movement_frame, MOVEMENT_REPORT_COLUMNS)

    frame = inventory.copy()
    if report_name == "Deleted Inventory Report":
        frame = frame.sort_values(["deleted_at", "id"], ascending=[False, False], na_position="last")
        return _select_columns(frame, DELETED_REPORT_COLUMNS)
    if report_name == "Reorder Report":
        flag = frame["stock_status"].isin(["Out of Stock", "Low Stock"])
        flag = flag | frame["reorder_quantity"].fillna(0).gt(0)
        frame = frame.loc[flag]
        frame = frame.sort_values(
            ["stock_status", "quantity_in_stock", "item_name"], na_position="last"
        )
    elif report_name == "Expiry Report":
        frame = frame.loc[frame["expiry_status"].isin(["Expired", "Expiring Soon"])]
        frame = frame.sort_values(["expiry_date", "item_name"], na_position="last")
    elif report_name == "Inventory Valuation Report":
        frame = frame.loc[frame["unit_price"].notna() & frame["quantity_in_stock"].notna()]
        frame = frame.sort_values("inventory_value", ascending=False, na_position="last")
    elif report_name == "Data Quality Report":
        frame = frame.loc[frame["data_quality_issues"].astype(str).str.len().gt(0)]
        frame = frame.sort_values(["item_name", "inventory_code"])
    else:
        frame = frame.sort_values(["item_name", "expiry_date"], na_position="last")
    return _select_columns(frame, INVENTORY_REPORT_COLUMNS)


def _select_columns(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    available = [column for column in columns if column in frame.columns]
    return frame.loc[:, available].copy()


def display_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[column]):
            result[column] = result[column].dt.date
    return result.rename(columns=FRIENDLY_COLUMNS)


def csv_bytes(frame: pd.DataFrame) -> bytes:
    return display_frame(frame).to_csv(index=False).encode("utf-8-sig")


def create_excel_report(
    report_title: str,
    frame: pd.DataFrame,
    summary: Mapping[str, Any],
    currency_symbol: str = "₹",
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    data_sheet = workbook.create_sheet("Report Data")

    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:D2")
    title_cell = summary_sheet["A1"]
    title_cell.value = report_title
    title_cell.font = Font(name="Arial", size=20, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center", horizontal="left")
    for row in summary_sheet["A1:D2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)

    summary_sheet["A4"] = "Generated"
    summary_sheet["B4"] = datetime.now().strftime("%d %b %Y, %I:%M %p")
    summary_sheet["A4"].font = Font(name="Arial", bold=True, color=GREY)
    summary_sheet["B4"].font = Font(name="Arial", color=NAVY)

    row_number = 6
    for label, value in summary.items():
        summary_sheet.cell(row_number, 1, label)
        summary_sheet.cell(row_number, 2, value)
        summary_sheet.cell(row_number, 1).font = Font(name="Arial", bold=True, color=NAVY)
        summary_sheet.cell(row_number, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        summary_sheet.cell(row_number, 2).font = Font(name="Arial", bold=True, color=NAVY)
        if "Value" in label:
            summary_sheet.cell(row_number, 2).number_format = f'"{currency_symbol}"#,##0.00'
        row_number += 1

    summary_sheet.cell(row_number + 1, 1, "Report Rows")
    summary_sheet.cell(row_number + 1, 2, len(frame))
    summary_sheet.cell(row_number + 1, 1).font = Font(name="Arial", bold=True, color=NAVY)
    summary_sheet.cell(row_number + 1, 1).fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    summary_sheet.cell(row_number + 1, 2).font = Font(name="Arial", bold=True, color=TEAL)
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 22
    summary_sheet.column_dimensions["C"].width = 4
    summary_sheet.column_dimensions["D"].width = 4

    exported = display_frame(frame)
    headers = list(exported.columns)
    for column_index, header in enumerate(headers, 1):
        cell = data_sheet.cell(1, column_index, header)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, values in enumerate(exported.itertuples(index=False, name=None), 2):
        for column_index, value in enumerate(values, 1):
            cell = data_sheet.cell(row_index, column_index)
            if value is None or (isinstance(value, float) and pd.isna(value)):
                cell.value = None
            elif isinstance(value, pd.Timestamp):
                cell.value = value.to_pydatetime()
            else:
                cell.value = value
            cell.font = Font(name="Arial", size=9, color=NAVY)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")

    thin = Side(style="thin", color="D9E2EA")
    for row in data_sheet.iter_rows(min_row=1, max_row=max(1, data_sheet.max_row)):
        for cell in row:
            cell.border = Border(bottom=thin)

    for column_index, header in enumerate(headers, 1):
        letter = get_column_letter(column_index)
        values = [str(header)] + [
            str(data_sheet.cell(row, column_index).value or "")
            for row in range(2, min(data_sheet.max_row, 80) + 1)
        ]
        width = min(max(len(value) for value in values) + 2, 34)
        data_sheet.column_dimensions[letter].width = max(width, 11)
        if any(term in header for term in ("Price", "Value", "Cost", "Profit")):
            for row in range(2, data_sheet.max_row + 1):
                data_sheet.cell(row, column_index).number_format = f'"{currency_symbol}"#,##0.00'
        elif "Date" in header or header == "Recorded At":
            for row in range(2, data_sheet.max_row + 1):
                data_sheet.cell(row, column_index).number_format = "dd-mmm-yyyy"
        elif "Quantity" in header or header in {
            "Reorder Level",
            "Total Test Effective",
            "Total Test Avail",
            "Total Test Done",
        }:
            for row in range(2, data_sheet.max_row + 1):
                data_sheet.cell(row, column_index).number_format = "#,##0.00"

    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    data_sheet.sheet_view.showGridLines = False
    data_sheet.row_dimensions[1].height = 34
    if headers and data_sheet.max_row >= 2:
        table = Table(displayName="InventoryReport", ref=data_sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        data_sheet.add_table(table)

    status_columns = [
        index
        for index, header in enumerate(headers, 1)
        if header in {"Stock Status", "Expiry Status", "Overall Status"}
    ]
    for column_index in status_columns:
        letter = get_column_letter(column_index)
        cell_range = f"{letter}2:{letter}{max(2, data_sheet.max_row)}"
        data_sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Expired"'], fill=PatternFill("solid", fgColor="FCE8E8")),
        )
        data_sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"Expiring Soon"'], fill=PatternFill("solid", fgColor="FFF4D6")),
        )
        data_sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"In Stock"'], fill=PatternFill("solid", fgColor="E4F4EC")),
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Import"
    headers = [
        "Item Name",
        "Brand",
        "Category",
        "Test Volume",
        "Description",
        "Create Date",
        "Unit Price",
        "Total Test Effective",
        "Total Test Avail",
        "Total Test Done",
        "Selling Price Per Test",
        "Quantity In Stock",
        "Reorder Level",
        "Reorder Quantity",
        "Opened Date",
        "Expiry Date",
        "Finish Date",
        "Reminder Days",
        "Reorder Due Date",
        "Batch / Lot No.",
        "Supplier",
        "Storage Location",
        "Notes",
    ]
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(1, index, header)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 3, 14), 24)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.row_dimensions[1].height = 34
    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "Inventory Import Instructions"
    instructions["A1"].font = Font(name="Arial", size=18, bold=True, color=WHITE)
    instructions["A1"].fill = PatternFill("solid", fgColor=NAVY)
    instructions["A3"] = "Required field"
    instructions["B3"] = "Item Name"
    instructions["A4"] = "Automatic code"
    instructions["B4"] = "Inventory ID and Inventory Code are generated by the app."
    instructions["A5"] = "Create Date"
    instructions["B5"] = "Optional for imports; records added in the app receive today's date."
    instructions["A6"] = "Dates"
    instructions["B6"] = "Use YYYY-MM-DD or DD-MM-YYYY."
    instructions["A7"] = "Deleted records"
    instructions["B7"] = "Deleted records are excluded from imports and remain in the Deleted tab."
    instructions["A8"] = "Replace mode"
    instructions["B8"] = "Creates a new inventory list after confirmation; make a backup first."
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 70
    for row in range(3, 9):
        instructions.cell(row, 1).font = Font(name="Arial", bold=True, color=NAVY)
        instructions.cell(row, 2).font = Font(name="Arial", color=NAVY)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
